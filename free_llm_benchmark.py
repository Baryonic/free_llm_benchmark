import requests #type: ignore
import datetime
import csv
import concurrent.futures
from deep_translator import GoogleTranslator
import os
import re
import shutil
import time
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter



# (in cmd) setx OPENROUTER_API_KEY "api-key"
API_KEY = os.environ.get("OPENROUTER_API_KEY")

# API endpoints and headers
API_URL = "https://openrouter.ai/api/v1/chat/completions"
MODEL_LIST_URL = "https://openrouter.ai/api/v1/models"
GEMINI_MODEL = "google/gemini-2.5-pro-exp-03-25"
headers = {
    "Authorization": f"Bearer {API_KEY}",
    "Content-Type": "application/json"
}

# Configurable delay and retry settings
REQUEST_DELAY = .2  # seconds between API requests
MAX_RETRIES = 3    # number of retries for 429 or 503 errors
INITIAL_RETRY_BACKOFF = 5 # seconds to wait for the first retry

# Global variables for tracking
failed_questions = []
failed_files = []  # Track failed files with reasons
successful_questions = []
small_files = []  # Track files that are too small
blacklisted_models = set()  # Track blacklisted models

# Create required directories if they don't exist
HTML_DIR = os.path.join(os.getcwd(), "html")
HTML_FAILED_DIR = os.path.join(os.getcwd(), "html_failed")
XCELL_DIR = os.path.join(os.getcwd(), "xcell")
XCELL_FAILED_DIR = os.path.join(os.getcwd(), "xcell_failed")

for directory in [HTML_DIR, HTML_FAILED_DIR, XCELL_DIR, XCELL_FAILED_DIR]:
    if not os.path.exists(directory):
        try:
            os.makedirs(directory)
            print(f"\033[92mCreated directory: {directory}\033[0m")
        except Exception as e:
            print(f"\033[31mError creating directory {directory}: {str(e)}\033[0m")
            if directory == HTML_DIR:
                HTML_DIR = "html"
                if not os.path.exists(HTML_DIR):
                    os.makedirs(HTML_DIR)
            elif directory == HTML_FAILED_DIR:
                HTML_FAILED_DIR = "html_failed"
                if not os.path.exists(HTML_FAILED_DIR):
                    os.makedirs(HTML_FAILED_DIR)
            elif directory == XCELL_DIR:
                XCELL_DIR = "xcell"
                if not os.path.exists(XCELL_DIR):
                    os.makedirs(XCELL_DIR)
            elif directory == XCELL_FAILED_DIR:
                XCELL_FAILED_DIR = "xcell_failed"
                if not os.path.exists(XCELL_FAILED_DIR):
                    os.makedirs(XCELL_FAILED_DIR)

def check_repeated_content(content, threshold=200):
    """
    Check if content contains the same message repeated multiple times.
    Returns True if content appears to be repeated, False otherwise.
    """
    # Split content into words
    words = content.split()
    if not words:
        return False
    
    # Count occurrences of each word
    word_counts = {}
    for word in words:
        word_counts[word] = word_counts.get(word, 0) + 1
    
    # Check if any word appears too many times
    for count in word_counts.values():
        if count > threshold:
            return True
    
    return False

def load_blacklist():
    """
    Load blacklisted model IDs from blacklist.csv
    Returns a set of blacklisted model IDs
    """
    blacklist_file = "blacklist.csv"
    blacklisted = set()
    
    try:
        if os.path.exists(blacklist_file):
            with open(blacklist_file, "r", encoding="utf-8") as f:
                for line in f:
                    model_id = line.strip()
                    if model_id:  # Skip empty lines
                        blacklisted.add(model_id)
            print(f"\033[92mLoaded {len(blacklisted)} blacklisted models\033[0m")
        else:
            print(f"\033[93mNo blacklist file found at {blacklist_file}\033[0m")
    except Exception as e:
        print(f"\033[31mError loading blacklist: {str(e)}\033[0m")
    
    return blacklisted

def load_free_models():
    """
    Fetch available models from the API, filtering only for free models where pricing is 0.
    Returns a dictionary keyed by model IDs containing model name and context length.
    """
    try:
        response = requests.get(MODEL_LIST_URL, headers=headers)
        response.raise_for_status()
        data = response.json()

        free_models = {}
        if not data or not isinstance(data.get("data", []), list):
            print("No models available or invalid API response.")
            return {}

        for item in data.get("data", []):
            model_id = item.get("id", "Unknown ID")
            
            # Skip blacklisted models
            if model_id in blacklisted_models:
                continue
                
            name = item.get("name", model_id)
            context_length = item.get("context_length", 100)  # default if missing
            pricing = item.get("pricing", {})
            prompt_cost = pricing.get("prompt")
            completion_cost = pricing.get("completion")

            # Convert pricing values to float for proper comparison.
            try:
                prompt_cost_val = float(prompt_cost) if prompt_cost is not None else None
                completion_cost_val = float(completion_cost) if completion_cost is not None else None
            except (TypeError, ValueError):
                continue

            if model_id.endswith(":free") or (prompt_cost_val == 0 and completion_cost_val == 0):
                free_models[model_id] = {
                    "name": name,
                    "context_length": context_length
                }
        return free_models

    except requests.exceptions.RequestException as e:
        print(f"Error: {type(e).__name__} - {e}")
        return {}

def translate_text(text, target_language, model_id=None):
    """
    Translate text using deep-translator library.
    target_language should be either 'english' or 'spanish'
    """
    print(f"\n\033[93mTranslation Debug Info:\033[0m")
    print(f"\033[93m- Target Language: {target_language}\033[0m")
    print(f"\033[93m- Text to translate: {text}\033[0m")
    
    try:
        # Map target_language to language codes
        lang_map = {
            'english': 'en',
            'spanish': 'es'
        }
        
        # Determine source and target languages
        if target_language == 'english':
            source = 'es'
            target = 'en'
        else:  # spanish
            source = 'en'
            target = 'es'
        
        # Perform translation
        print(f"\033[93m- Sending translation request...\033[0m")
        translator = GoogleTranslator(source=source, target=target)
        translation = translator.translate(text)
        
        if translation:
            print(f"\033[92m- Translation successful:\033[0m")
            print(f"\033[92m  {translation}\033[0m")
            return translation
        else:
            print(f"\033[31m- Translation failed: No translation received\033[0m")
            return None
            
    except Exception as e:
        print(f"\033[31m- Translation failed: Unexpected Error\033[0m")
        print(f"\033[31m- Error Type: {type(e).__name__}\033[0m")
        print(f"\033[31m- Error Message: {str(e)}\033[0m")
        return None

def query_model(model_id, prompt, max_tokens):
    """
    Sends a query to the OpenRouter API for the given model using the provided prompt and max_tokens.
    Returns a tuple containing the processed response text and the raw response data.
    Implements retry logic for 429 and 503 errors with exponential backoff.
    """
    payload = {
        "messages": [{  
            "role": "user",
            "content": f"{prompt}." 
        }],
        "model": model_id,
        "max_tokens": max_tokens
    }
    attempt = 0
    while attempt <= MAX_RETRIES:
        print(f"[{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Sending request to {model_id} (attempt {attempt+1})")
        try:
            response = requests.post(API_URL, headers=headers, json=payload)
            
            if response.status_code in [429, 503]:
                wait_time = INITIAL_RETRY_BACKOFF * (2 ** attempt)
                print(f"\033[31mReceived {response.status_code} error. Waiting {wait_time} seconds before retry...\033[0m")
                time.sleep(wait_time)
                attempt += 1
                continue

            response.raise_for_status()
            try:
                data = response.json()
            except requests.exceptions.JSONDecodeError:
                return f"API Request Error: Invalid JSON response from model {model_id}. Response text: {response.text}", {}

            if data and "choices" in data:
                choices = data.get("choices", [])
                if choices and isinstance(choices, list):
                    # Get the message content
                    message = choices[0].get("message", {})
                    processed_response = message.get("content", "No response.")
                    # Check if the response contains error indicators or chain of thought
                    if any(indicator in processed_response.lower() for indicator in ["error:", "i apologize", "i'm sorry", "i cannot", "i don't", "i'm not sure", "i'm unable"]):
                        # Keep the error message or chain of thought as is
                        return processed_response, data
                    elif not processed_response.strip():
                        return "Error: Empty response received.", data
                else:
                    processed_response = "Error: Invalid API structure."
            else:
                processed_response = "Error: No valid response received."
            return processed_response, data
        except requests.exceptions.RequestException as e:
            if hasattr(e, 'response') and e.response is not None and e.response.status_code in [429, 503]:
                wait_time = INITIAL_RETRY_BACKOFF * (2 ** attempt)
                print(f"\033[31mCaught {e.response.status_code} error in exception. Waiting {wait_time} seconds before retry...\033[0m")
                time.sleep(wait_time)
                attempt += 1
                continue
            return f"API Request Error: {str(e)}", {}
        break
    return f"API Request Error: Failed after {MAX_RETRIES+1} attempts", {}

def create_excel_report_for_prompt(original_spanish_prompt, english_prompt, results, timestamp, safe_prompt):
    """
    Creates an Excel report with the same information as the HTML report.
    Returns the filename if successful, None if failed.
    """
    filename = os.path.join(XCELL_DIR, f"{safe_prompt}_{timestamp}.xlsx")
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Model Responses"
        
        # Set up styles
        header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        headers = [
            "Model", "Model ID", "Prompt Tokens", "Completion Tokens", 
            "Total Tokens", "Characters", "Chars/Token", 
            "English Response", "Spanish Translation"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Write data
        row = 2
        for result in results:
            model_name = result["model_name"]
            model_id = result.get("model_id", "Unknown ID")
            
            # Process token information
            tokens = result.get("tokens", {})
            if isinstance(tokens, dict):
                prompt_tokens = tokens.get("prompt_tokens", "N/A")
                completion_tokens = tokens.get("completion_tokens", "N/A")
                total_tokens = tokens.get("total_tokens", "N/A")
            else:
                prompt_tokens = "N/A"
                completion_tokens = "N/A"
                total_tokens = tokens
            
            # Calculate character counts and efficiency
            english_response = result.get("english_response", "No response")
            char_count = len(english_response) if english_response != "No response" else 0
            
            if isinstance(total_tokens, (int, float)) and total_tokens > 0:
                efficiency = f"{char_count/total_tokens:.2f}"
            else:
                efficiency = "N/A"
            
            # Write row data
            row_data = [
                model_name, model_id, prompt_tokens, completion_tokens,
                total_tokens, char_count, efficiency,
                english_response, result.get("spanish_response", "Translation failed")
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # Save the workbook
        wb.save(filename)
        return filename
        
    except Exception as e:
        print(f"\033[31mError creating Excel report: {str(e)}\033[0m")
        return None

def safe_move_file(filename, reason, is_excel=False):
    """
    Safely move a file to the failed directory, handling any permission errors.
    Returns True if move was successful, False otherwise.
    """
    try:
        if os.path.exists(filename):
            # Get the base filename
            base_filename = os.path.basename(filename)
            # Create the destination path
            dest_dir = XCELL_FAILED_DIR if is_excel else HTML_FAILED_DIR
            dest_path = os.path.join(dest_dir, base_filename)
            # Move the file
            shutil.move(filename, dest_path)
            failed_files.append((dest_path, reason))
            return True
    except Exception as e:
        print(f"\033[31mError moving file {filename}: {str(e)}\033[0m")
    return False

def create_html_report_for_prompt(original_spanish_prompt, english_prompt, results):
    """
    Creates an HTML report with the original Spanish prompt, English translation,
    and responses from all models in both languages.
    """
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    safe_prompt = "".join(c for c in original_spanish_prompt if c.isalnum() or c in (' ', '_')).strip().replace(" ", "_")[:20]
    html_filename = os.path.join(HTML_DIR, f"{safe_prompt}_{timestamp}.html")

    html_content = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Model Comparison Report</title>
    <link rel="stylesheet" href="../css/estilo.css" media="all">
</head>
<body>
    <h1>Original Spanish Question:</h1>
    <p>{original_spanish_prompt}</p>
    <h2>English Translation (by Gemini):</h2>
    <p>{english_prompt}</p>
    <table>
        <thead>
            <tr>
                <th>Model</th>
                <th>Request Time</th>
                <th>Start Time</th> <!-- New column -->
                <th>End Time</th> <!-- New column -->
                <th>Duration (s)</th> <!-- New column -->
                <th>Token Usage</th>
                <th>Characters</th>
                <th>Chars/Token</th>
                <th>Timestamps</th> <!-- New column for timestamps -->
                <th>English Response</th>
                <th>Spanish Translation (by Google Translate)</th>
            </tr>
        </thead>
        <tbody>
    """

    for result in results:
        model_name = result["model_name"]
        model_id = result.get("model_id", "Unknown ID")
        request_time = result.get("request_time", "N/A")
        start_time = result.get("start_time", "N/A")  # New field
        end_time = result.get("end_time", "N/A")  # New field
        duration = result.get("duration", "N/A")  # New field

        # Process token information
        tokens = result.get("tokens", {})
        if isinstance(tokens, dict):
            prompt_tokens = tokens.get("prompt_tokens", "N/A")
            completion_tokens = tokens.get("completion_tokens", "N/A")
            total_tokens = tokens.get("total_tokens", "N/A")
            token_info = f'<div class="token-info">Prompt: {prompt_tokens}<br>Completion: {completion_tokens}<br>Total: {total_tokens}</div>'
        else:
            token_info = f'<div class="token-info">Total: {tokens}</div>'
            total_tokens = tokens

        # Calculate character counts and efficiency
        english_response = result.get("english_response", "No response")
        char_count = len(english_response) if english_response != "No response" else 0

        if isinstance(total_tokens, (int, float)) and total_tokens > 0:
            efficiency = f"{char_count/total_tokens:.2f}"
        else:
            efficiency = "N/A"

        char_info = f'<div class="char-info">{char_count:,}</div>'  
        efficiency_info = f'<div class="efficiency">{efficiency}</div>'

        # Format English response
        if english_response:
            formatted_response = english_response.replace("\n", "<br>")
            if any(indicator in english_response.lower() for indicator in ["error:", "i apologize", "i'm sorry", "i cannot", "i don't", "i'm not sure", "i'm unable"]):
                english_response = f'<div class="chain-of-thought">{formatted_response}</div>'
            elif english_response.startswith("Error:"):
                english_response = f'<div class="error-message">{formatted_response}</div>'
            else:
                english_response = formatted_response
        else:
            english_response = '<div class="error-message">No response</div>'
            
        # Format Spanish response
        spanish_response = result.get("spanish_response", "Translation failed")
        if spanish_response:
            formatted_response = spanish_response.replace("\n", "<br>")
            if any(indicator in spanish_response.lower() for indicator in ["error:", "lo siento", "no puedo", "no estoy seguro", "no soy capaz"]):
                spanish_response = f'<div class="chain-of-thought">{formatted_response}</div>'
            elif spanish_response.startswith("Error:"):
                spanish_response = f'<div class="error-message">{formatted_response}</div>'
            else:
                spanish_response = formatted_response
        else:
            spanish_response = '<div class="error-message">Translation failed</div>'
        
        html_content += f"""<tr>
            <td>{model_name}<br><span class=\"model-id\">{model_id}</span></td>
            <td>{request_time}</td>
            <td>{start_time}</td> <!-- New field -->
            <td>{end_time}</td> <!-- New field -->
            <td>{duration if isinstance(duration, (int, float)) else 'N/A'}</td> <!-- Fixed formatting -->
            <td>{token_info}</td>
            <td>{char_info}</td>
            <td>{efficiency_info}</td>
            <td>{start_time} - {end_time}</td> <!-- New column for timestamps -->
            <td class=\"response-cell\"><div class=\"english-response\">{english_response}</div></td>
            <td class=\"response-cell\"><div class=\"spanish-response\">{spanish_response}</div></td>
        </tr>
        """

    html_content += """
        </tbody>
    </table>
    <div class="footer">
        <p>Note: All Spanish translations were performed by Google Translate</p>
    </div>
    <script>
        // Synchronize scrolling between English and Spanish responses
        document.addEventListener('DOMContentLoaded', function() {
            const rows = document.querySelectorAll('tr');
            rows.forEach(row => {
                const englishCell = row.querySelector('.english-response');
                const spanishCell = row.querySelector('.spanish-response');
                
                if (englishCell && spanishCell) {
                    englishCell.addEventListener('scroll', function() {
                        spanishCell.scrollTop = this.scrollTop;
                    });
                    
                    spanishCell.addEventListener('scroll', function() {
                        englishCell.scrollTop = this.scrollTop;
                    });
                }
            });
        });
    </script>
</body>
</html>"""

    # Write the HTML file
    try:
        with open(html_filename, "w", encoding="utf-8") as file:
            file.write(html_content)
    except Exception as e:
        print(f"\033[31mError writing HTML file {html_filename}: {str(e)}\033[0m")
        return None

    # Check file size
    try:
        file_size = os.path.getsize(html_filename)
        if file_size < 50 * 1024:  # 50KB in bytes
            if safe_move_file(html_filename, f"file too small ({file_size/1024:.1f}KB)"):
                small_files.append((html_filename, file_size))
                print(f"\033[31mMoved HTML file {html_filename} to html_failed due to small size ({file_size/1024:.1f}KB)\033[0m")
            return None

        # If HTML is valid, create Excel report
        excel_filename = create_excel_report_for_prompt(original_spanish_prompt, english_prompt, results, timestamp, safe_prompt)
        if excel_filename:
            print(f"Excel report saved as '{excel_filename}'")
        else:
            print(f"\033[31mFailed to create Excel report\033[0m")
            if safe_move_file(html_filename, "failed to create Excel report"):
                print(f"\033[31mMoved HTML file {html_filename} to html_failed due to Excel creation failure\033[0m")
            return None

        print(f"HTML report saved as '{html_filename}' ({file_size/1024:.1f}KB)")
        return html_filename
    except Exception as e:
        print(f"\033[31mError processing file {html_filename}: {str(e)}\033[0m")
        return None

def process_model_response(args):
    model_id, details, max_tokens, english_question = args
    print(f"\n\033[96m{'='*80}\033[0m")
    print(f"\033[96mProcessing Model: {details['name']}\033[0m")
    print(f"\033[96m{'='*80}\033[0m")
    try:
        # Record start time
        start_time = datetime.datetime.now()
        print(f"\033[93mStart Time: {start_time.strftime('%Y-%m-%d %H:%M:%S')}\033[0m")

        # Get English response
        print(f"\n\033[93mSending English prompt to {details['name']}:\033[0m")
        print(f"\033[93mPrompt: {english_question}\033[0m")
        english_response, raw_data = query_model(model_id, english_question, max_tokens)
        print(f"\n\033[92mReceived English response from {details['name']}:\033[0m")
        print(f"\033[92mResponse: {english_response}\033[0m")

        # Record end time and calculate duration
        end_time = datetime.datetime.now()
        duration = (end_time - start_time).total_seconds()
        print(f"\033[93mEnd Time: {end_time.strftime('%Y-%m-%d %H:%M:%S')}\033[0m")
        print(f"\033[93mDuration: {duration:.2f} seconds\033[0m")

        # Add delay between requests for rate limiting
        print(f"\033[90mWaiting {REQUEST_DELAY} seconds before next request...\033[0m")
        time.sleep(REQUEST_DELAY)

        # Check if response is an error message
        is_error = any(error in english_response.lower() for error in [
            "error:", "no valid response received", "no response", "error processing response"
        ])

        # Only translate if not an error message
        if not is_error:
            print(f"\n\033[95mTranslating response from {details['name']} back to Spanish using Gemini...\033[0m")
            print(f"\033[95mEnglish text to translate: {english_response}\033[0m")
            spanish_response = translate_text(english_response, "spanish", GEMINI_MODEL)

            if not spanish_response:
                print(f"\033[31mWarning: Translation failed for {details['name']}\033[0m")
                print(f"\033[31m- Model: {details['name']}\033[0m")
                print(f"\033[31m- Original English response: {english_response}\033[0m")
                print(f"\033[31m- Translation model used: {GEMINI_MODEL}\033[0m")
                spanish_response = "Translation failed"
            else:
                print(f"\033[95mSpanish translation (by Gemini): {spanish_response}\033[0m")
        else:
            print(f"\n\033[93mSkipping translation for error message\033[0m")
            spanish_response = english_response  # Use the same error message in Spanish

        # Process token information
        tokens = {}
        if raw_data and "usage" in raw_data:
            usage = raw_data["usage"]
            tokens = {
                "prompt_tokens": usage.get("prompt_tokens", "N/A"),
                "completion_tokens": usage.get("completion_tokens", "N/A"),
                "total_tokens": usage.get("total_tokens", "N/A")
            }
            print(f"\033[94mToken usage: {tokens}\033[0m")
        else:
            tokens = "N/A"
            print(f"\033[94mToken usage: N/A\033[0m")

        request_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        print(f"\033[96m{'='*80}\033[0m")
        return {
            "model_name": details["name"],
            "model_id": model_id,
            "tokens": tokens,
            "english_response": english_response,
            "spanish_response": spanish_response,
            "translated_by": "Google Translate",
            "request_time": request_time,  # <-- add this line
            "start_time": start_time.strftime('%Y-%m-%d %H:%M:%S'),
            "end_time": end_time.strftime('%Y-%m-%d %H:%M:%S'),
            "duration": duration
        }
    except Exception as e:
        print(f"\033[31mError processing {details['name']}: {str(e)}\033[0m")
        print(f"\033[31m- Error Type: {type(e).__name__}\033[0m")
        print(f"\033[31m- Error Message: {str(e)}\033[0m")
        return {
            "model_name": details["name"],
            "model_id": model_id,
            "tokens": "N/A",
            "english_response": "Error processing response",
            "spanish_response": "Error processing response",
            "translated_by": "Google Translate",
            "start_time": "N/A",
            "end_time": "N/A",
            "duration": "N/A"
        }

def process_question(question):
    """
    Process a single question through the translation and response pipeline
    """
    print(f"\n\033[95m{'='*80}\033[0m")
    print(f"\n\033[95mProcessing Spanish Question:\033[0m")
    print(f"\033[95mQuestion: {question}\033[0m")
    print(f"\033[95m{'='*80}\033[0m")
    
    # Translate to English
    print("\n\033[93mStep 1: Translating question to English...\033[0m")
    english_question = translate_text(question, "english")
    if not english_question:
        print("\033[31mFailed to translate question to English. Skipping.\033[0m")
        failed_questions.append((question, "translation failed"))
        return None
    
    print(f"\n\033[92mStep 1 Complete: English Translation:\033[0m")
    print(f"\033[92mTranslation: {english_question}\033[0m")
    
    # Get free models
    print("\n\033[93mStep 2: Loading available free models...\033[0m")
    free_models = load_free_models()
    if not free_models:
        print("No free models available. Exiting.")
        failed_questions.append((question, "no free models available"))
        return None
    
    print(f"\n\033[92mStep 2 Complete: Found {len(free_models)} free models\033[0m")
    for model_id, details in free_models.items():
        print(f"\033[92m- {details['name']}\033[0m")

    # Prepare arguments for parallel execution
    print("\n\033[93mStep 3: Preparing for parallel processing...\033[0m")
    model_args = []
    for model_id, details in free_models.items():
        try:
            max_tokens = int(details["context_length"] // 2)
        except Exception:
            max_tokens = 50
        model_args.append((model_id, details, max_tokens, english_question))
        print(f"\033[92m- {details['name']}: max_tokens={max_tokens}\033[0m")

    print("\n\033[93mStep 4: Starting parallel processing of models...\033[0m")
    # Use ThreadPoolExecutor for parallel requests
    results = []
    with concurrent.futures.ThreadPoolExecutor() as executor:
        future_to_model = {executor.submit(process_model_response, args): args for args in model_args}
        for future in concurrent.futures.as_completed(future_to_model):
            result = future.result()
            results.append(result)
            print(f"\n\033[92mCompleted processing for {result['model_name']}\033[0m")

    print("\n\033[93mStep 5: Generating HTML report...\033[0m")
    # Create HTML report
    report_file = create_html_report_for_prompt(question, english_question, results)
    
    if report_file is None:
        print(f"\033[31mFailed to generate valid report for question: {question}\033[0m")
        failed_questions.append((question, "invalid HTML report"))
        return None
    
    print(f"\n\033[92mStep 5 Complete: Report generated as {report_file}\033[0m")
    print(f"\033[95m{'='*80}\033[0m")
    successful_questions.append(question)
    return True

def process_pending_questions():
    """
    Reads pending questions from 'preguntas_pendientes.csv' and processes them
    with translation workflow.
    """
    print(f"\n\033[94m{'='*80}\033[0m")
    print("\033[94mStarting Question Processing Pipeline\033[0m")
    print(f"\033[94m{'='*80}\033[0m")
    
    # Load blacklist at the start
    global blacklisted_models
    blacklisted_models = load_blacklist()
    
    print("\n\033[93mStep 1: Reading pending questions file...\033[0m")
    pending_file = "preguntas_pendientes.csv"
    resolved_file = "preguntas_resueltas.csv"

    try:
        with open(pending_file, "r", encoding="utf-8") as pf:
            pending_lines = pf.readlines()
        print(f"\033[92mSuccessfully read {pending_file}\033[0m")
    except FileNotFoundError:
        print(f"\033[31mError: File {pending_file} not found.\033[0m")
        return

    pending_questions = [line.strip() for line in pending_lines if line.strip()]
    print(f"\033[92mFound {len(pending_questions)} pending questions\033[0m")

    if not pending_questions:
        print("\n\033[93mNo pending questions found. Requesting new question...\033[0m")
        new_question = input("\033[31mNo pending questions found. \033[92mPlease enter a new question for the models (or type 'exit'/'quit' to end):\033[0m ").strip().lower()
        if new_question in ['exit', 'quit']:
            print("\033[92mExiting program as requested.\033[0m")
            return
        elif new_question:
            pending_questions = [new_question]
            print("\033[92mNew question added to processing queue\033[0m")
        else:
            print("\033[31mNo question provided. Exiting.\033[0m")
            return

    remaining_questions = []

    print("\n\033[93mStep 2: Processing questions...\033[0m")
    for i, question in enumerate(pending_questions, 1):
        if question:
            print(f"\n\033[94mProcessing question {i} of {len(pending_questions)}:\033[0m")
            print(f"\033[94mQuestion: {question}\033[0m")
            if process_question(question):
                print(f"\n\033[93mStep 3: Updating resolved questions file...\033[0m")
                # Append the processed question to the resolved file
                with open(resolved_file, "a", encoding="utf-8") as rf:
                    rf.write(question + "\n")
                print(f"\033[92mQuestion added to resolved file\033[0m")
            else:
                print(f"\033[31mFailed to process question. Adding to remaining questions.\033[0m")
                remaining_questions.append(question)
        else:
            remaining_questions.append(question)

    print("\n\033[93mStep 4: Updating pending questions file...\033[0m")
    # Update pending file with remaining questions
    with open(pending_file, "w", encoding="utf-8") as pf:
        for q in remaining_questions:
            pf.write(q + "\n")
    print(f"\033[92mUpdated pending questions file with {len(remaining_questions)} remaining questions\033[0m")

    # Generate final report
    print(f"\n\033[94m{'='*20}\033[0m")
    print(f"\033[93mProcessing Summary Report\033[0m")
    print(f"\033[94m{'='*20}\033[0m")
    
    print("\n\033[92mSuccessfully Processed Questions:\033[0m")
    print(f"Total: {len(successful_questions)}")
    for q in successful_questions:
        print(f"- {q}")
    
    print("\n\033[31mFailed Questions (Returned to Pending):\033[0m")
    print(f"Total: {len(failed_questions)}")
    for q, reason in failed_questions:
        print(f"- {q} (Reason: {reason})")
    
    print("\n\033[33mFailed Files:\033[0m")
    print(f"Total: {len(failed_files)}")
    for f, reason in failed_files:
        file_size = os.path.getsize(f)
        print(f"- {f} ({file_size/1024:.1f}KB) - Reason: {reason}")
    
    print("\n\033[33mSmall Files (Under 50KB):\033[0m")
    print(f"Total: {len(small_files)}")
    for f, size in small_files:
        print(f"- {f} ({size/1024:.1f}KB)")
    
    print("\n\033[35mBlacklisted Models:\033[0m")
    print(f"Total: {len(blacklisted_models)}")
    for model_id in sorted(blacklisted_models):
        print(f"- {model_id}")
    
    print(f"\n\033[94m{'='*80}\033[0m")
    print("\033[94mQuestion Processing Pipeline Complete\033[0m")
    print(f"\033[94m{'='*80}\033[0m")

if __name__ == "__main__":
    print("\033[94mrunning \033[92mFREE LLM BENCHMARK \033[94mby \033[95mKEYDAY ELECTRONICS SOFTWARE \033[94mand \033[95mRUMI EXPLORA")
    print(f"\033[94mExecuting from: \033[93m{os.path.abspath(__file__)}")
    process_pending_questions()